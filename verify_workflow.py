import time
import requests
import json
from pathlib import Path
import os
import signal
import subprocess

BASE_URL = "http://localhost:8000/api/v1"

def print_step(msg):
    print(f"\n{'='*50}\n[STEP] {msg}\n{'='*50}")

def wait_for_server():
    for _ in range(30):
        try:
            r = requests.get("http://localhost:8000/health")
            if r.status_code == 200:
                print("Server is up!")
                return
        except:
            pass
        time.sleep(1)
    raise Exception("Server failed to start")

def main():
    print_step("Starting Server")
    server_proc = subprocess.Popen(["python", "-m", "uvicorn", "backend.app.main:app", "--port", "8000"])
    
    try:
        wait_for_server()
        
        # 0. Setup: Ensure employee exists and an attendance record exists
        print_step("0. Setup Data")
        emps = requests.get(f"{BASE_URL}/employees").json()
        if not emps:
            print("Creating dummy employee")
            emp = requests.post(f"{BASE_URL}/employees", json={
                "employee_code": "E999",
                "name": "Test Emp",
                "monthly_salary": 30000,
                "working_days_per_month": 26
            }).json()
        else:
            emp = emps[0]
            
        emp_id = emp["id"]
        emp_code = emp["employee_code"]
        work_date = "2026-07-20"
        
        # We need an attendance record for them to show in reports
        # But report_service fetches AttendanceRepository.list_for_date. 
        # If no attendance data, we can ingest some dummy data
        csv_data = f"Employee ID,Employee Name,Date,Check In,Check Out\n{emp_code},Test Emp,2026-07-20,09:00,17:00"
        with open("dummy_attendance.csv", "w") as f:
            f.write(csv_data)
            
        with open("dummy_attendance.csv", "rb") as f:
            requests.post(f"{BASE_URL}/attendance/upload", files={"file": f})
            
        # 1. Add Sick Leave annotation
        print_step("1. Add a Sick Leave annotation")
        r = requests.put(f"{BASE_URL}/annotations/{emp_id}/{work_date}", json={
            "employee_id": emp_id,
            "work_date": work_date,
            "annotation_type": "Sick Leave",
            "notes": "Feeling unwell"
        })
        r.raise_for_status()
        ann = r.json()
        ann_id = ann["id"]
        print("Created annotation:", ann)
        
        # 2. Confirm it is saved
        print_step("2. Confirm it is saved in the database")
        r = requests.get(f"{BASE_URL}/annotations", params={"work_date": work_date})
        r.raise_for_status()
        anns = r.json()
        assert len(anns) >= 1
        assert any(a["id"] == ann_id and a["annotation_type"] == "Sick Leave" for a in anns)
        print("Annotation successfully verified in DB.")
        
        # 4. Edit it
        print_step("4. Edit annotation")
        r = requests.put(f"{BASE_URL}/annotations/{emp_id}/{work_date}", json={
            "employee_id": emp_id,
            "work_date": work_date,
            "annotation_type": "Medical Leave",
            "notes": "Doctor appointment"
        })
        r.raise_for_status()
        
        r = requests.get(f"{BASE_URL}/annotations", params={"work_date": work_date})
        anns = r.json()
        assert any(a["id"] == ann_id and a["annotation_type"] == "Medical Leave" for a in anns)
        print("Annotation successfully edited to Medical Leave.")
        
        # 5. Delete it
        print_step("5. Delete annotation")
        r = requests.delete(f"{BASE_URL}/annotations/{ann_id}")
        r.raise_for_status()
        
        r = requests.get(f"{BASE_URL}/annotations", params={"work_date": work_date})
        anns = r.json()
        assert not any(a["id"] == ann_id for a in anns)
        print("Annotation successfully deleted.")
        
        # Now add one back for the report generation test
        print_step("Re-add annotation for reports test")
        r = requests.put(f"{BASE_URL}/annotations/{emp_id}/{work_date}", json={
            "employee_id": emp_id,
            "work_date": work_date,
            "annotation_type": "Work From Home"
        })
        
        # 6. Generate Reports
        print_step("6. Generate Reports (CSV, Excel, PDF)")
        reports = {}
        for fmt in ["csv", "excel", "pdf"]:
            r = requests.post(f"{BASE_URL}/reports/generate", json={
                "report_type": "daily_summary",
                "format": fmt,
                "work_date": work_date
            })
            r.raise_for_status()
            res = r.json()
            print(f"Generated {fmt} report:", res["filename"])
            reports[fmt] = res["filename"]
            
        print_step("Verifying Generated Reports Content")
        import pandas as pd
        import openpyxl
        
        # Verify CSV
        print("Checking CSV...")
        csv_path = Path("reports") / reports["csv"]
        df_csv = pd.read_csv(csv_path)
        print("CSV Columns:", df_csv.columns.tolist())
        row = df_csv[df_csv["Employee ID"] == emp_code].iloc[0]
        reason = row.get("Reason", "")
        print(f"Reason in CSV: '{reason}'")
        assert reason == "Work From Home", "CSV missing annotation"
        hours = row["Work Hours"]
        print(f"Work Hours in CSV: {hours} (type: {type(hours)})")
        assert isinstance(hours, (int, float)) or str(hours).replace('.', '', 1).isdigit(), "CSV hours not numeric"
        
        # Verify Excel
        print("Checking Excel...")
        xl_path = Path("reports") / reports["excel"]
        wb = openpyxl.load_workbook(xl_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        reason_idx = headers.index("Reason")
        hours_idx = headers.index("Work Hours")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == emp_code: # assuming Employee ID is col 0
                xl_reason = row[reason_idx]
                xl_hours = row[hours_idx]
                print(f"Reason in Excel: '{xl_reason}'")
                print(f"Work Hours in Excel: '{xl_hours}'")
                assert xl_reason == "Work From Home", "Excel missing annotation"
                assert "h" in str(xl_hours) or "m" in str(xl_hours), f"Excel hours not human readable: {xl_hours}"
                break
        
        print_step("All verifications passed!")
        
    finally:
        server_proc.terminate()
        server_proc.wait()

if __name__ == "__main__":
    main()
