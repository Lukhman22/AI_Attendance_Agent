"""Regression tests for CSV / XLS / XLSX attendance uploads and content detection."""

from __future__ import annotations

from datetime import date, time
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook
from xlwt import Workbook as XlsWorkbook

from backend.app.attendance.csv_reader import _detect_file_format, read_attendance_file
from backend.app.core.exceptions import ApplicationError


FLAT_HEADERS = [
    "Employee ID",
    "Employee Name",
    "Department",
    "Date",
    "Check-In Time",
    "Check-Out Time",
    "Work Duration",
    "Break Duration",
    "Overtime",
    "Attendance Status",
]
FLAT_ROW = ["E001", "John Doe", "Engineering", "2026-07-14", "09:00", "17:00", "7.25", "1.00", "0.00", "Present"]


def _csv_bytes() -> bytes:
    header = ",".join(FLAT_HEADERS)
    row = ",".join(FLAT_ROW)
    return f"{header}\n{row}\n".encode("utf-8")


def _xlsx_bytes(path: Path) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"
    sheet.append(FLAT_HEADERS)
    sheet.append(FLAT_ROW)
    workbook.save(path)
    return path.read_bytes()


def _xls_bytes(path: Path) -> bytes:
    workbook = XlsWorkbook()
    sheet = workbook.add_sheet("Attendance")
    for col, value in enumerate(FLAT_HEADERS):
        sheet.write(0, col, value)
    for col, value in enumerate(FLAT_ROW):
        sheet.write(1, col, value)
    workbook.save(str(path))
    return path.read_bytes()


def _monthly_xlsx_bytes(path: Path) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"
    sheet.append(["Monthly Attendance Report - July 2026"])
    sheet.append([])
    sheet.append(["Employee Code", "E001", "Employee Name", "John Doe", "Department", "Engineering"])
    sheet.append(["", 1, 2, 3, 4, 5, 6, 7])
    sheet.append(["IN", "09:00", "09:00", "", "09:00", "", "09:00", ""])
    sheet.append(["OUT", "18:00", "17:00", "", "19:30", "", "", ""])
    sheet.append(["WORK", "8.00", "7.00", "", "8.00", "", "", ""])
    sheet.append(["Break", "1.00", "1.00", "", "1.00", "", "0.50", ""])
    sheet.append(["OT", "0", "0", "", "1.50", "", "0", ""])
    sheet.append(["Status", "P", "P", "WO", "P", "LV", "P", "A"])
    workbook.save(path)
    return path.read_bytes()


def _monthly_xls_bytes(path: Path) -> bytes:
    workbook = XlsWorkbook()
    sheet = workbook.add_sheet("Attendance")
    sheet.write(0, 0, "Monthly Attendance Report - July 2026")
    sheet.write(2, 0, "Employee Code")
    sheet.write(2, 1, "E002")
    sheet.write(2, 2, "Employee Name")
    sheet.write(2, 3, "Jane Smith")
    for col, day in enumerate(range(1, 8), start=1):
        sheet.write(3, col, day)
    metrics = [
        ("IN", ["09:00", "09:00", "09:00", "09:00", "", "09:00", ""]),
        ("OUT", ["18:00", "18:00", "18:00", "18:00", "", "18:00", ""]),
        ("WORK", ["8.00", "8.00", "8.00", "8.00", "", "8.00", ""]),
        ("Break", ["1.00", "1.00", "1.00", "1.00", "", "1.00", ""]),
        ("OT", ["0", "0", "0", "0", "", "0", ""]),
        ("Status", ["P", "P", "P", "P", "WO", "P", "HL"]),
    ]
    for row_idx, (label, values) in enumerate(metrics, start=4):
        sheet.write(row_idx, 0, label)
        for col_idx, value in enumerate(values, start=1):
            sheet.write(row_idx, col_idx, value)
    workbook.save(str(path))
    return path.read_bytes()


def test_detect_format_from_content_magic(tmp_path: Path):
    csv_data = _csv_bytes()
    xlsx_data = _xlsx_bytes(tmp_path / "a.xlsx")
    xls_data = _xls_bytes(tmp_path / "a.xls")

    assert _detect_file_format(csv_data, "attendance.csv") == "csv"
    assert _detect_file_format(xlsx_data, "attendance.xlsx") == "xlsx"
    assert _detect_file_format(xls_data, "attendance.xls") == "xls"


def test_detect_format_ignores_wrong_extension(tmp_path: Path):
    csv_data = _csv_bytes()
    xlsx_data = _xlsx_bytes(tmp_path / "real.xlsx")
    xls_data = _xls_bytes(tmp_path / "real.xls")

    # Content wins over misleading extensions
    assert _detect_file_format(csv_data, "report.xls") == "csv"
    assert _detect_file_format(xlsx_data, "monthperformance.csv") == "xlsx"
    assert _detect_file_format(xls_data, "export.xlsx") == "xls"


def test_parse_flat_csv():
    rows = read_attendance_file(BytesIO(_csv_bytes()), "attendance.csv")
    assert len(rows) == 1
    assert rows[0].employee_code == "E001"
    assert rows[0].work_duration_hours == Decimal("7.25")
    assert rows[0].work_date == date(2026, 7, 14)


def test_parse_flat_xlsx(tmp_path: Path):
    data = _xlsx_bytes(tmp_path / "flat.xlsx")
    rows = read_attendance_file(BytesIO(data), "flat.xlsx")
    assert len(rows) == 1
    assert rows[0].employee_code == "E001"
    assert rows[0].check_in == time(9, 0)
    assert rows[0].work_duration_hours == Decimal("7.25")


def test_parse_flat_xls(tmp_path: Path):
    data = _xls_bytes(tmp_path / "flat.xls")
    rows = read_attendance_file(BytesIO(data), "flat.xls")
    assert len(rows) == 1
    assert rows[0].employee_code == "E001"
    assert rows[0].employee_name == "John Doe"
    assert rows[0].work_duration_hours == Decimal("7.25")
    assert rows[0].break_duration_hours == Decimal("1.00")


def test_parse_xls_with_csv_extension(tmp_path: Path):
    """Production quirk: legacy .xls exported with a .csv filename."""
    data = _xls_bytes(tmp_path / "monthperformance.xls")
    rows = read_attendance_file(BytesIO(data), "monthperformance14072026163400.csv")
    assert len(rows) == 1
    assert rows[0].employee_code == "E001"


def test_parse_xlsx_with_csv_extension(tmp_path: Path):
    data = _xlsx_bytes(tmp_path / "modern.xlsx")
    rows = read_attendance_file(BytesIO(data), "attendance_export.csv")
    assert len(rows) == 1
    assert rows[0].employee_code == "E001"


def test_parse_csv_with_xls_extension():
    rows = read_attendance_file(BytesIO(_csv_bytes()), "attendance.xls")
    assert len(rows) == 1
    assert rows[0].work_duration_hours == Decimal("7.25")


def test_parse_monthly_xlsx_block(tmp_path: Path):
    data = _monthly_xlsx_bytes(tmp_path / "monthly.xlsx")
    rows = read_attendance_file(BytesIO(data), "monthly.xlsx")
    assert len(rows) == 7
    assert {r.employee_code for r in rows} == {"E001"}
    assert all(r.work_date.month == 7 and r.work_date.year == 2026 for r in rows)


def test_parse_monthly_xls_block(tmp_path: Path):
    data = _monthly_xls_bytes(tmp_path / "monthly.xls")
    rows = read_attendance_file(BytesIO(data), "monthly.xls")
    assert len(rows) == 7
    assert rows[0].employee_code == "E002"
    assert rows[0].employee_name == "Jane Smith"


def test_parse_production_company_monthly_xls():
    """Regression: real Celutron monthperformance .xls (Empcode + weekday row layout)."""
    path = Path("sample_data/monthperformance_june_2026.xls")
    if not path.exists():
        pytest.skip("production sample not present")
    rows = read_attendance_file(path.open("rb"), path.name)
    assert len(rows) >= 100
    codes = {r.employee_code for r in rows}
    assert "K6k031" in codes or "K6K016" in codes
    assert all(r.work_date.year == 2026 and r.work_date.month == 6 for r in rows)
    azeem = [r for r in rows if r.employee_code == "K6k031"]
    assert azeem
    assert azeem[0].employee_name == "Azeem"
    assert azeem[0].work_duration_hours is not None
    # WORK values like 08:20 must become decimal hours
    assert azeem[0].work_duration_hours > 0


def test_parse_monthly_with_weekday_row_and_empcode(tmp_path: Path):
    """Company layout: Empcode/Name meta, day numbers, Mon-Sun row, then IN..Status."""
    from xlwt import Workbook as XlsWorkbook

    path = tmp_path / "company_block.xls"
    wb = XlsWorkbook()
    sheet = wb.add_sheet("Sheet1")
    sheet.write(0, 0, "Dept. Name")
    sheet.write(0, 2, "Engineering")
    sheet.write(0, 26, "Report Month")
    sheet.write(0, 29, "July-2026")
    sheet.write(1, 0, "Empcode")
    sheet.write(1, 2, "E100")
    sheet.write(1, 5, "Name")
    sheet.write(1, 7, "Samir")
    sheet.write(1, 12, "Present")
    sheet.write(1, 14, "5")
    for col, day in enumerate(range(1, 8), start=1):
        sheet.write(2, col, str(day))
    weekdays = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    for col, day_name in enumerate(weekdays, start=1):
        sheet.write(3, col, day_name)
    sheet.write(4, 0, "IN")
    sheet.write(5, 0, "OUT")
    sheet.write(6, 0, "WORK")
    sheet.write(7, 0, "Break")
    sheet.write(8, 0, "OT")
    sheet.write(9, 0, "Status")
    for col in range(1, 8):
        sheet.write(4, col, "09:00")
        sheet.write(5, col, "18:00")
        sheet.write(6, col, "08:00")
        sheet.write(7, col, "01:00")
        sheet.write(8, col, "00:00")
        sheet.write(9, col, "P" if col != 7 else "WO")
    wb.save(str(path))

    rows = read_attendance_file(path.open("rb"), path.name)
    assert len(rows) == 7
    assert rows[0].employee_code == "E100"
    assert rows[0].employee_name == "Samir"
    assert rows[0].department == "Engineering"
    assert rows[0].work_date == date(2026, 7, 1)
    assert rows[0].work_duration_hours == Decimal("8.00")
    assert rows[6].status == "wo"


def test_parse_error_logs_and_raises_application_error(tmp_path: Path, caplog):
    garbage = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"not-a-real-workbook"
    with caplog.at_level("ERROR"), pytest.raises(ApplicationError) as exc_info:
        read_attendance_file(BytesIO(garbage), "broken.xls")
    assert exc_info.value.code == "attendance_file_parse_error"
    assert exc_info.value.message == "Unable to parse attendance file"
    assert "Unable to parse attendance file" in caplog.text
