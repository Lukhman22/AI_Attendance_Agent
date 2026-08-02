import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

def format_excel_report(
    workbook: Workbook,
    sheet,
    title: str,
    report_period: str,
    summary_data: dict[str, str],
    headers: list[str],
    rows: list[list[Any]],
    filename: Path
) -> Path:
    # 1. Set Workbook metadata
    workbook.properties.title = "AI Attendance Agent Report"
    workbook.properties.author = "AI Attendance Agent"
    workbook.properties.company = "Your Company"

    # 2. Add Title Section
    sheet.append(["AI Attendance Agent"])
    sheet.append([title])
    if report_period:
        sheet.append([report_period.replace(': ', ':\n')])
    now = datetime.datetime.now()
    sheet.append([f"Generated On:\n{now.strftime('%d %B %Y')}\n{now.strftime('%I:%M %p')}"])
    sheet.append([]) # Blank row
    
    current_row = sheet.max_row
    
    if summary_data:
        summary_items = [f"{k}:\n{v}" for k, v in summary_data.items()]
        sheet.append(summary_items)
        sheet.append([])
        current_row = sheet.max_row

    # Style Title Section
    title_font = Font(name="Calibri", size=16, bold=True, color="1F2937")
    subtitle_font = Font(name="Calibri", size=14, bold=True, color="4B5563")
    meta_font = Font(name="Calibri", size=10, color="4B5563")
    summary_font = Font(name="Calibri", size=10, bold=True, color="1F2937")
    
    sheet["A1"].font = title_font
    sheet["A2"].font = subtitle_font
    if report_period:
        sheet["A3"].font = meta_font
        sheet["A3"].alignment = Alignment(wrap_text=True, vertical="top")
        sheet["B3"].font = meta_font
        sheet["B3"].alignment = Alignment(wrap_text=True, vertical="top", horizontal="right")
        sheet["B3"].value = sheet["A4"].value # Move generated on to B3
        sheet["A4"].value = ""
    else:
        sheet["A3"].font = meta_font
        sheet["A3"].alignment = Alignment(wrap_text=True, vertical="top")

    if summary_data:
        for col_idx in range(1, len(summary_items) + 1):
            cell = sheet.cell(row=current_row - 1, column=col_idx)
            cell.font = summary_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    header_row_idx = current_row + 1
    if headers:
        sheet.append(headers)
    for row in rows:
        sheet.append(row)

    # 4. Styling Variables
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    # Define alignments
    align_left = Alignment(horizontal="left", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Determine column formatting by header name
    col_formats = {}
    for col_idx, h in enumerate(headers, start=1):
        h_lower = h.lower()
        if "name" in h_lower or "reason" in h_lower or "department" in h_lower:
            align = align_left
        elif "id" in h_lower or "status" in h_lower or "time" in h_lower or "hour" in h_lower or "%" in h_lower or "percentage" in h_lower:
            align = align_center
        elif "salary" in h_lower or "deduction" in h_lower or "amount" in h_lower or "currency" in h_lower or "total" in h_lower:
            align = align_right
        else:
            align = align_left # default fallback
        
        num_format = "General"
        if "salary" in h_lower or "deduction" in h_lower or "amount" in h_lower:
            num_format = '"₹"#,##0.00'
        elif "%" in h_lower or "percentage" in h_lower:
            num_format = '0.00%'
        elif "date" in h_lower:
            num_format = 'dd/mm/yyyy'
        elif "time" in h_lower and "overtime" not in h_lower:
            num_format = 'hh:mm:ss'
            
        col_formats[col_idx] = {"align": align, "format": num_format}

    # 5. Apply Styles, Borders, and Formats
    total_rows = len(rows) + header_row_idx
    for row_idx in range(header_row_idx, total_rows + 1):
        is_header = (row_idx == header_row_idx)
        sheet.row_dimensions[row_idx].height = 20 if is_header else 15
        
        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            
            if is_header:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = Font(name="Calibri", size=11)
                fmt = col_formats.get(col_idx)
                if fmt:
                    cell.alignment = fmt["align"]
                    cell.number_format = fmt["format"]

    # 6. Adjust Column Widths
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(header_row_idx, total_rows + 1):
            val_str = str(sheet.cell(row=row_idx, column=col_idx).value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        sheet.column_dimensions[col_letter].width = max(max_len + 2, 10)

    # 7. AutoFilter & Freeze Panes
    if headers:
        sheet.freeze_panes = f"A{header_row_idx + 1}"
        
        # Only add Table if we have data rows (Excel tables require at least 1 data row)
        if rows:
            from openpyxl.worksheet.table import TableColumn
            ref = f"A{header_row_idx}:{get_column_letter(len(headers))}{total_rows}"
            tab = Table(displayName="ReportData", ref=ref)
            style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            # Explicitly define columns to prevent XML corruption missing column metadata
            tab.tableColumns = [TableColumn(id=i+1, name=str(h)) for i, h in enumerate(headers)]
            sheet.add_table(tab)
        else:
            # Fallback to just autofilter if no data rows
            ref = f"A{header_row_idx}:{get_column_letter(len(headers))}{header_row_idx}"
            sheet.auto_filter.ref = ref

    # 8. Conditional Formatting
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.styles.differential import DifferentialStyle
    
    # Status colors
    green_fill = PatternFill(bgColor="C6EFCE", fill_type="solid")
    green_text = Font(color="006100")
    red_fill = PatternFill(bgColor="FFC7CE", fill_type="solid")
    red_text = Font(color="9C0006")
    yellow_fill = PatternFill(bgColor="FFEB9C", fill_type="solid")
    yellow_text = Font(color="9C5700")
    blue_fill = PatternFill(bgColor="BDD7EE", fill_type="solid")
    blue_text = Font(color="203764")
    gray_fill = PatternFill(bgColor="D9D9D9", fill_type="solid")
    gray_text = Font(color="595959")
    
    status_rules = [
        ("Present", green_fill, green_text),
        ("Absent", red_fill, red_text),
        ("Leave", yellow_fill, yellow_text),
        ("Holiday", blue_fill, blue_text),
        ("Weekly Off", gray_fill, gray_text)
    ]
    
    for col_idx, h in enumerate(headers, start=1):
        h_lower = h.lower()
        col_letter = get_column_letter(col_idx)
        data_range = f"{col_letter}{header_row_idx + 1}:{col_letter}{total_rows}"
        
        if h_lower == "status":
            for text, fill, font in status_rules:
                rule = CellIsRule(operator='equal', formula=[f'"{text}"'], fill=fill, font=font)
                sheet.conditional_formatting.add(data_range, rule)
                
        if "deduction" in h_lower:
            # Zero -> green text
            sheet.conditional_formatting.add(data_range, CellIsRule(operator='equal', formula=['0'], font=green_text))
            # Positive -> red text
            sheet.conditional_formatting.add(data_range, CellIsRule(operator='greaterThan', formula=['0'], font=red_text))

    # 9. Print Settings
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0  # 0 means auto
    sheet.print_title_rows = f"{header_row_idx}:{header_row_idx}"
    sheet.page_margins.left = 0.5
    sheet.page_margins.right = 0.5
    sheet.page_margins.top = 0.5
    sheet.page_margins.bottom = 0.5

    workbook.save(filename)
    return filename
